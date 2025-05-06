import imaplib
import email
from email.header import decode_header
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

# CONFIGURAÇÕES
EMAIL = "victoraraujo.zeba@gmail.com"       # <-- Substitua
SENHA = "voto ldho tddk rfgv"                # <-- Substitua
SERVIDOR_IMAP = "imap.gmail.com"
PLANILHA = "Orcamentos.xlsx"

# Conecta ao email
mail = imaplib.IMAP4_SSL(SERVIDOR_IMAP)
mail.login(EMAIL, SENHA)
mail.select("inbox")

# Procura emails não lidos
status, mensagens = mail.search(None, '(UNSEEN)')
if status != "OK":
    print("Nenhum email novo.")
    exit()

# Abre a planilha
wb = openpyxl.load_workbook(PLANILHA)
mes_atual = datetime.now().strftime("%m-%Y")

# Cria nova aba do mês a partir da aba "Modelo"
if mes_atual not in wb.sheetnames:
    if "Modelo" in wb.sheetnames:
        aba_modelo = wb["Modelo"]
        ws = wb.copy_worksheet(aba_modelo)
        ws.title = mes_atual

        # Centraliza conteúdo das células
        for col in ws.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    else:
        print("❌ Aba 'Modelo' não encontrada. Abortando criação da nova aba.")
        exit()
else:
    ws = wb[mes_atual]

# Lista de palavras-chave para filtrar
palavras_chave = ["orçamento", "orcamento", "cotação", "cotacao", "pedido", "compra", "solicitação"]

# Processa os e-mails
for num in mensagens[0].split():
    # NÃO marca como lido automaticamente
    status, dados = mail.fetch(num, "(BODY.PEEK[])")
    raw_email = dados[0][1]
    msg = email.message_from_bytes(raw_email)

    assunto_raw = msg["Subject"]
    assunto, cod = decode_header(assunto_raw)[0]
    if isinstance(assunto, bytes):
        try:
            assunto = assunto.decode(cod or "utf-8")
        except:
            assunto = assunto.decode("latin-1")

    if not any(palavra in assunto.lower() for palavra in palavras_chave):
        continue  # ignora e-mail que não tem palavras-chave

    remetente = msg["From"]
    agora = datetime.now()
    data_formatada = agora.strftime("%d/%m/%Y")
    hora_formatada = agora.strftime("%H:%M")


    # Pega o corpo do e-mail
    corpo = ""
    if msg.is_multipart():
        for parte in msg.walk():
            if parte.get_content_type() == "text/plain":
                try:
                    corpo = parte.get_payload(decode=True).decode()
                except UnicodeDecodeError:
                    try:
                        corpo = parte.get_payload(decode=True).decode("latin-1")
                    except:
                        corpo = parte.get_payload(decode=True).decode("iso-8859-1")
                break
    else:
        try:
            corpo = msg.get_payload(decode=True).decode()
        except UnicodeDecodeError:
            try:
                corpo = msg.get_payload(decode=True).decode("latin-1")
            except:
                corpo = msg.get_payload(decode=True).decode("iso-8859-1")

    from_email = email.utils.parseaddr(remetente)[1]

    # Adiciona na planilha
   # Encontra a primeira linha vazia a partir da linha 2
    # 🧮 Número do orçamento: conta linhas já preenchidas (ignora o cabeçalho)
    numero_orcamento = ws.max_row + 1  # Incrementa para evitar pular linhas

    # 👤 Extrai apenas o nome do cliente (antes do <...>)
    nome_cliente = remetente.split("<")[0].strip()

    # 🧾 Monta a linha completa no modelo da planilha
    nova_linha = [
        numero_orcamento,
        nome_cliente,
        from_email,
        assunto,
        corpo.strip(),
        data_formatada,
        hora_formatada,
        "Email",  # Origem
        "", "", "",  # ML Preço, Link, Prazo
        "", "", "",  # Amazon Preço, Link, Prazo
        "", "", "",  # Shopee Preço, Link, Prazo
    ]

    # Adiciona a nova linha na planilha
    ws.append(nova_linha)

    # Centraliza o conteúdo das células da nova linha
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(horizontal="center", vertical="center")



    print(f"📩 Novo pedido de: {remetente} | Assunto: {assunto}")

    # ✅ Agora marca como lido
    mail.store(num, '+FLAGS', '\\Seen')

# Salva a planilha
wb.save(PLANILHA)
mail.logout()
