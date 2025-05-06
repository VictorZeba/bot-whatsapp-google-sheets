import openpyxl
print("Módulo openpyxl carregado.")

import requests
print("Módulo requests carregado.")

from bs4 import BeautifulSoup
print("Módulo BeautifulSoup carregado.")
from datetime import datetime

print("Iniciando o bot...")

# Caminho para a planilha
planilha_path = "Orcamentos.xlsx"

try:
    # Carrega a planilha
    wb = openpyxl.load_workbook(planilha_path)
    print("Planilha carregada com sucesso.")

    # Pega a aba do mês atual
    mes_atual = datetime.now().strftime("%m-%Y")
    if mes_atual not in wb.sheetnames:
        print(f"Aba '{mes_atual}' não encontrada.")
        exit()  # Encerra o script se a aba não existir
    else:
        ws = wb[mes_atual]
        print(f"Aba '{mes_atual}' carregada.")

    # Itera pelas linhas da planilha
    for row in ws.iter_rows(min_row=2):
        item_cell = row[4]  # Coluna E – "Itens do Pedido"
        ml_preco = row[8]  # Coluna I – ML Preço

        # Pula se já tem preço preenchido ou item em branco
        if ml_preco.value or not item_cell.value:
            continue

        termo_busca = item_cell.value.strip()
        print(f"🔍 Buscando: {termo_busca}")

        try:
            # Cria a URL de busca no Mercado Livre
            url = f"https://lista.mercadolivre.com.br/{termo_busca.replace(' ', '-')}"
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(url, headers=headers)

            # Verifica o status da requisição
            if response.status_code != 200:
                print(f"❌ Falha na requisição: Status Code {response.status_code}")
                row[8].value = "Erro"
                row[9].value = "Erro"
                row[10].value = "Erro"
                continue

            soup = BeautifulSoup(response.text, "html.parser")

            # Pega o primeiro produto listado
            resultado = soup.find("li", class_="ui-search-layout__item")
            if not resultado:
                print("❌ Nenhum resultado encontrado.")
                row[8].value = "Não encontrado"
                row[9].value = "Não encontrado"
                row[10].value = "Não encontrado"
                continue

            # Extrai informações do produto
            link_tag = resultado.find("a")
            titulo_tag = resultado.find("h2")
            preco_principal_tag = resultado.find("span", class_="andes-money-amount__fraction")
            preco_centavos_tag = resultado.find("span", class_="andes-money-amount__cents")

            # Verifica se os elementos existem
            link = link_tag["href"] if link_tag and "href" in link_tag.attrs else "Link não encontrado"
            titulo = titulo_tag.text.strip() if titulo_tag else "Título não encontrado"
            preco_inteiro = preco_principal_tag.text.strip() if preco_principal_tag else "0"
            preco_fracao = preco_centavos_tag.text.strip() if preco_centavos_tag else "00"

            # Remove pontos de milhar e substitui vírgula por ponto
            preco_final = f"{preco_inteiro}.{preco_fracao}".replace(".", "").replace(",", ".")

            # Converte para float
            try:
                preco = float(preco_final)  # Converte para float
            except ValueError:
                print(f"⚠️ Erro ao converter preço: {preco_final}")
                preco = 0.0  # Define como 0.0 em caso de erro

            # Exibe o preço formatado
            print(f"✅ Preço encontrado: R${preco:.2f}")

            # Extrai prazo de entrega
            prazo_tag = resultado.find("p", class_="ui-search-item__shipping")
            prazo = prazo_tag.text.strip() if prazo_tag else "Prazo não disponível"

            # Preenche os dados na planilha
            row[8].value = preco  # Coluna I – ML Preço
            row[9].value = link  # Coluna J – ML Link
            row[10].value = prazo  # Coluna K – ML Prazo
            print(f"✅ Produto encontrado: {titulo} – R${preco:.2f} – {prazo}")

        except Exception as e:
            print(f"⚠️ Erro na busca: {e}")
            row[8].value = "Erro"
            row[9].value = "Erro"
            row[10].value = "Erro"

    # Salva a planilha (apenas uma vez ao final)
    wb.save(planilha_path)
    print("📁 Planilha atualizada com os preços do Mercado Livre.")

except Exception as e:
    print(f"❌ Erro geral: {e}")